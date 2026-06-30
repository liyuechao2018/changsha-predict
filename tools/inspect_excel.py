from __future__ import annotations

import sys
from pathlib import Path

import openpyxl


def main() -> None:
    path = Path(sys.argv[1])
    wb = openpyxl.load_workbook(path, data_only=True)
    print("sheets:", wb.sheetnames)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n[{sheet_name}] rows={ws.max_row} cols={ws.max_column}")
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 8), values_only=True):
            print(tuple(row))


if __name__ == "__main__":
    main()
