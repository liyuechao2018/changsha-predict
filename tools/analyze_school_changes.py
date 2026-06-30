from __future__ import annotations

import re
import sys
from dataclasses import dataclass
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from database import ADMISSIONS_2025  # noqa: E402


ALIASES = {
    "长郡中学": "长郡",
    "雅礼中学": "雅礼",
    "师大附中": "师大附中",
    "湖南师大附中": "师大附中",
    "附中": "师大附中",
    "长沙市一中": "市一中",
    "一中": "市一中",
    "麓山": "麓山国际",
    "麓山国际": "麓山国际",
    "南雅中学": "南雅",
    "南雅": "南雅",
    "明德中学": "明德",
    "明德": "明德",
    "周南中学": "周南",
    "周南": "周南",
    "长沙市实验中学": "市实验",
    "实验": "市实验",
    "市实验": "市实验",
    "师大梅溪湖": "师梅",
    "师梅": "师梅",
    "长郡双语": "长郡双语",
    "雅礼洋湖": "雅礼洋湖",
    "长郡梅溪湖": "长梅",
    "长梅": "长梅",
    "周南梅溪湖": "周梅",
    "周梅": "周梅",
    "长郡滨江": "长郡滨江",
    "长沙市六中": "市六中",
    "六中": "市六中",
    "6中": "市六中",
    "市六中": "市六中",
    "附中博才": "附中博才湘江",
    "附中博才实验(湘江)": "附中博才湘江",
    "附中博才湘江": "附中博才湘江",
    "长沙外国语": "长外",
    "长外": "长外",
    "望城一中": "望城一中",
    "望城区一中": "望城一中",
    "麓山梅溪湖": "麓山梅溪湖",
    "麓梅": "麓山梅溪湖",
    "麓山滨江": "麓山滨江",
    "明德华兴": "明德华兴",
    "长郡湘府": "长郡湘府",
    "长铁一中": "长铁一中",
    "长沙铁路一中": "长铁一中",
    "铁一": "长铁一中",
    "周南实验": "周南实验",
    "长沙市十五中": "市十五中",
    "十五中": "市十五中",
    "15中": "市十五中",
    "市十五中": "市十五中",
    "长沙市十一中": "市十一中城南路",
    "11中": "市十一中城南路",
    "十一中(文化班)": "市十一中城南路",
    "市十一中城南路": "市十一中城南路",
    "南雅梅溪湖": "南雅梅溪湖",
    "地质中学": "地质",
    "地质": "地质",
    "雅礼书院": "雅礼书院",
    "雅礼书院中学": "雅礼书院",
    "长沙市二十一中": "二十一中",
    "21中": "二十一中",
    "二十一中": "二十一中",
    "科大附中": "科大附中",
    "长郡斑马湖": "长郡斑马湖",
    "长郡智谷": "长郡智谷",
    "一中城南": "一中城南",
    "田家炳实验": "田家炳实验",
    "田家炳": "田家炳实验",
    "市十一中融城": "市十一中融城",
    "明德雨花": "明德雨花",
    "明德雨花实验": "明德雨花",
    "稻田中学": "稻田",
    "稻田": "稻田",
    "雷锋学校": "雷锋",
    "雷锋": "雷锋",
    "附中雨花": "附中雨花",
    "师大附中雨花学校": "附中雨花",
    "师大附中雨花": "附中雨花",
    "麓山外国语": "麓山外国语",
    "一中广雅": "一中广雅",
    "湖大附中": "湖大附中",
    "长大附中": "长大附中",
    "长沙大学附属中学": "长大附中",
    "雅礼中南附属": "雅礼中南附属",
    "岳麓实验": "岳麓实验",
    "东雅中学": "东雅",
    "东雅": "东雅",
}


@dataclass
class SchoolRecord:
    year: int
    name: str
    canonical: str
    score: float | None
    rank: int | None
    rank_label: str
    plan: int | None = None


def canonicalize(name: object) -> str:
    text = str(name or "").strip()
    text = re.sub(r"\s+", "", text)
    text = text.replace("长沙市", "")
    return ALIASES.get(text, text)


def parse_rank(value: object) -> tuple[int | None, str]:
    if value is None:
        return None, ""
    if isinstance(value, (int, float)):
        return int(round(float(value))), str(int(round(float(value))))
    text = str(value).strip().replace(",", "").replace("，", "")
    numbers = [int(item) for item in re.findall(r"\d+", text)]
    if not numbers:
        return None, text
    if len(numbers) >= 2 and "-" in text:
        return int(round((numbers[0] + numbers[1]) / 2)), text
    return numbers[-1], text


def load_2024(path: Path) -> list[SchoolRecord]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["2024第一批"]
    records: list[SchoolRecord] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[2]:
            continue
        rank, rank_label = parse_rank(row[5])
        records.append(
            SchoolRecord(
                year=2024,
                name=str(row[2]).strip(),
                canonical=canonicalize(row[2]),
                score=float(row[4]) if row[4] is not None else None,
                rank=rank,
                rank_label=rank_label,
                plan=int(row[3]) if isinstance(row[3], (int, float)) else None,
            )
        )
    return records


def load_2023(path: Path) -> list[SchoolRecord]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["2023"]
    records: list[SchoolRecord] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        try:
            score = float(row[1]) if row[1] is not None else None
        except (TypeError, ValueError):
            continue
        rank, rank_label = parse_rank(row[5])
        records.append(
            SchoolRecord(
                year=2023,
                name=str(row[0]).strip(),
                canonical=canonicalize(row[0]),
                score=score,
                rank=rank,
                rank_label=rank_label,
            )
        )
    return records


def load_2025() -> list[SchoolRecord]:
    return [
        SchoolRecord(
            year=2025,
            name=name,
            canonical=canonicalize(name),
            score=float(score),
            rank=int(rank_without_quota),
            rank_label=str(rank_without_quota),
        )
        for name, score, rank_without_quota, _rank_with_quota in ADMISSIONS_2025
    ]


def top(records: list[SchoolRecord], n: int = 40) -> list[SchoolRecord]:
    return sorted([r for r in records if r.rank is not None], key=lambda r: r.rank or 10**9)[:n]


def compare(previous: list[SchoolRecord], current: list[SchoolRecord], label: str) -> None:
    previous_map = {r.canonical: r for r in top(previous, 45)}
    current_top = top(current, 45)
    new_records = [r for r in current_top if r.canonical not in previous_map]
    print(f"\n## {label} 新进入头部参考范围")
    if not new_records:
        print("无")
        return
    for r in new_records:
        print(f"- {r.canonical}: {r.score:g}分, 约第 {r.rank:,} 名, 原始位次={r.rank_label}, 计划={r.plan or ''}")


def movement(previous: list[SchoolRecord], current: list[SchoolRecord], label: str) -> None:
    previous_map = {r.canonical: r for r in previous if r.rank is not None}
    current_map = {r.canonical: r for r in current if r.rank is not None}
    shared = []
    for name, now in current_map.items():
        old = previous_map.get(name)
        if old:
            shared.append((name, old.rank, now.rank, (now.rank or 0) - (old.rank or 0), old.score, now.score))
    shared.sort(key=lambda item: abs(item[3]), reverse=True)
    print(f"\n## {label} 共同学校位次变化较大的样本")
    for name, old_rank, now_rank, diff, old_score, now_score in shared[:12]:
        direction = "后移" if diff > 0 else "前移"
        print(f"- {name}: {old_rank:,} -> {now_rank:,}, {direction} {abs(diff):,} 名, 分数 {old_score:g}->{now_score:g}")


def print_top(records: list[SchoolRecord], label: str) -> None:
    print(f"\n## {label} 头部40所")
    for i, r in enumerate(top(records, 40), start=1):
        print(f"{i:02d}. {r.canonical} | {r.score:g}分 | 约{r.rank:,}名 | 原始={r.rank_label} | 计划={r.plan or ''}")


def main() -> None:
    path = Path(sys.argv[1])
    records_2023 = load_2023(path)
    records_2024 = load_2024(path)
    records_2025 = load_2025()

    print_top(records_2023, "2023")
    print_top(records_2024, "2024")
    print_top(records_2025, "2025")
    compare(records_2023, records_2024, "2023 -> 2024")
    compare(records_2024, records_2025, "2024 -> 2025")
    movement(records_2023, records_2024, "2023 -> 2024")
    movement(records_2024, records_2025, "2024 -> 2025")


if __name__ == "__main__":
    main()
